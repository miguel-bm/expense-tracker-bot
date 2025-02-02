from datetime import datetime
from io import BytesIO
from typing import Any, Literal

import xlrd
from openai import AsyncOpenAI
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import ContextTypes

from app.bot.utils import filter_message, send_typing_action
from app.models.expense import Expense
from app.models.income import Income
from app.models.movement import Movement
from app.storage.expenses.base import ExpenseStorageInterface
from app.storage.incomes.base import IncomeStorageInterface
from app.utils.logger import logger
from app.utils.movement_classifier.main import process_movements


def value_to_date(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    elif isinstance(value, float):
        raise ValueError("Float is not a valid date")
    elif isinstance(value, int):
        raise ValueError("Int is not a valid date")
    try:
        return datetime.strptime(value, "%d/%m/%Y")
    except ValueError:
        # Fall back to 2-digit year
        try:
            return datetime.strptime(value, "%d/%m/%y")
        except ValueError:
            raise ValueError(f"Invalid date format: {value}")


def read_eur_currency_amount(value: Any) -> float:
    value_str = str(value).strip()
    if value_str.endswith("EUR"):
        value_str = value_str[:-3]
    value_str = value_str.replace(".", "")
    value_str = value_str.replace(",", ".")
    return float(value_str)


def process_bbva(file_bytes: bytes) -> list[Movement]:
    wb = load_workbook(filename=BytesIO(file_bytes))
    sheet = wb.worksheets[0]
    movements = []

    # Start from row 5 and column B (which is index 1)
    row = 6
    while sheet.cell(row=row, column=2).value:  # Check if Fecha cell is not empty
        concept = str(sheet.cell(row=row, column=4).value)
        movement_str = str(sheet.cell(row=row, column=5).value)
        observations = str(sheet.cell(row=row, column=10).value)
        text = f"{concept}; {movement_str}; {observations}"
        date_ = value_to_date(sheet.cell(row=row, column=2).value)
        date_value = value_to_date(sheet.cell(row=row, column=3).value)
        movements.append(
            Movement(
                date_=date_,
                date_value=date_value,
                description=text,
                amount=float(sheet.cell(row=row, column=6).value or 0.0),  # type: ignore
                balance=float(sheet.cell(row=row, column=8).value or 0.0),  # type: ignore
            )
        )
        row += 1

    return movements


def process_caixabank(file_bytes: bytes) -> list[Movement]:
    wb: xlrd.book.Book = xlrd.open_workbook(file_contents=file_bytes)
    sheet: xlrd.sheet.Sheet = wb.sheet_by_index(0)
    movements = []
    for row_idx in range(3, sheet.nrows):  # Start from row 4
        if not sheet.cell_value(row_idx, 0):  # Empty first cell
            break

        # Convert Excel dates to Python dates
        date_ = xlrd.xldate.xldate_as_datetime(sheet.cell_value(row_idx, 0), 0).date()
        date_value = xlrd.xldate.xldate_as_datetime(
            sheet.cell_value(row_idx, 1), 0
        ).date()

        concept = str(sheet.cell_value(row_idx, 2) or "")
        description = str(sheet.cell_value(row_idx, 3) or "")
        text = f"{concept}; {description}"

        movements.append(
            Movement(
                date_=date_,
                date_value=date_value,
                description=text,
                amount=float(sheet.cell_value(row_idx, 4) or 0.0),
                balance=float(sheet.cell_value(row_idx, 5) or 0.0),
            )
        )
    return movements


def process_imagin(file_bytes: bytes) -> list[Movement]:
    wb = load_workbook(filename=BytesIO(file_bytes))
    sheet = wb.worksheets[0]
    movements = []

    row = 4
    while sheet.cell(row=row, column=1).value:
        description = str(sheet.cell(row=row, column=1).value)
        date_ = value_to_date(sheet.cell(row=row, column=2).value)
        amount = read_eur_currency_amount(sheet.cell(row=row, column=3).value)
        balance = read_eur_currency_amount(sheet.cell(row=row, column=4).value)
        movements.append(
            Movement(
                date_=date_,
                date_value=date_,
                description=description,
                amount=amount,
                balance=balance,
            )
        )
        row += 1

    return movements


def discriminate_format(
    file_bytes: bytes, file_extension: str
) -> Literal["bbva", "imagin"] | None:
    """Discriminate the format of the sheet"""
    if file_extension == "xls":
        return "imagin"
    elif file_extension == "xlsx":
        return "bbva"
    return None


def process_tabular_file(
    file_bytes: bytes, file_extension: str, file_name: str
) -> list[Movement]:
    """Process tabular files (XLS, XLSX) and return movements."""
    try:
        file_format = discriminate_format(file_bytes, file_extension)

        if file_format == "imagin":
            return process_caixabank(file_bytes)
        elif file_format == "bbva":
            if file_name.startswith("Movimientos "):
                return process_imagin(file_bytes)
            return process_bbva(file_bytes)
        else:
            raise ValueError("Unsupported file format")

    except Exception as e:
        logger.exception(f"Error processing file: {str(e)}")
        raise e


@send_typing_action
@filter_message
async def handle_file_message(
    update: Update, context: ContextTypes.DEFAULT_TYPE
) -> None:
    assert update.message and update.message.document
    user_mapping: dict[str, str] = context.bot_data["user_mapping"]
    expense_storage: ExpenseStorageInterface = context.bot_data["expense_storage"]
    income_storage: IncomeStorageInterface = context.bot_data["income_storage"]
    openai_client: AsyncOpenAI = context.bot_data["openai"]

    file = update.message.document
    from_user = update.message.from_user
    username = from_user.username if from_user and from_user.username else "unknown"
    username = user_mapping.get(username or "", username)

    logger.info(f"Received file from {username}: {file.file_name}")

    if file.file_name is None:
        await update.message.reply_text(
            "Sorry, I couldn't process this file because I couldn't get the file name."
        )
        return

    file_extension = file.file_name.lower().split(".")[-1]
    file_name = file.file_name
    file_content = await file.get_file()
    file_bytes = await file_content.download_as_bytearray()

    # Process the file
    if file_extension not in ["xls", "xlsx"]:
        await update.message.reply_text(
            "Sorry, I can only process CSV, XLS, or XLSX files at the moment."
        )
        return

    try:
        await update.message.reply_text("Processing file...")
        movements = process_tabular_file(file_bytes, file_extension, file_name)
        movements = sorted(movements, key=lambda x: x.min_date, reverse=False)
        new_expenses_or_incomes = await process_movements(
            movements, username, expense_storage, income_storage, openai_client
        )
    except Exception as e:
        logger.exception(f"Error processing file: {str(e)}")
        await update.message.reply_text(
            f"Sorry, I couldn't process this file because something went wrong: {str(e)}"
        )
        return

    logger.info(f"Processed {len(movements)} movements")
    num_new_expenses = sum(
        1 for expense in new_expenses_or_incomes if isinstance(expense, Expense)
    )
    num_new_incomes = sum(
        1 for income in new_expenses_or_incomes if isinstance(income, Income)
    )
    num_classified_as_other = sum(
        1 for e in new_expenses_or_incomes if e.category == ["other"]
    )
    await update.message.reply_text(
        f"I've processed the file and added the movements to your expenses and incomes. "
        f"There were {len(movements)} movements in the file, of which {num_new_expenses} were new expenses and {num_new_incomes} were new incomes. "
        f"There were {num_classified_as_other} movements that I could not classify and I advise you to classify them manually in the Google Sheet."
    )
    return
